import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def flatten_dict(d, parent_key='', sep='_'):
    """Recursively flatten a nested dictionary."""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        
        if isinstance(v, dict):
            if "text" in v and len(v) == 1:  # Special case for {key: {"text": value}}
                items.append((new_key, v["text"]))
            else:
                items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list) and all(isinstance(x, dict) for x in v):
            # Skip lists of dicts - they'll be handled separately as tables
            continue
        else:
            items.append((new_key, v))
    return dict(items)

def format_key_for_display(key):
    """Format a key for display in Excel."""
    # Convert snake_case or camelCase to Title Case With Spaces
    words = []
    current_word = ""
    
    for char in key:
        if char == '_':
            if current_word:
                words.append(current_word)
                current_word = ""
        elif char.isupper() and current_word and not current_word[-1].isupper():
            words.append(current_word)
            current_word = char
        else:
            current_word += char
            
    if current_word:
        words.append(current_word)
        
    return " ".join(word.capitalize() for word in words)

def json_to_excel(json_data, output_path, doc_type_code=None):
    """
    Convert JSON data to Excel format dynamically.
    
    Args:
        json_data (dict): The JSON data to convert
        output_path (str): Path to save the Excel file
        doc_type_code (str, optional): Document type code for sheet naming
    """
    # Handle case when json_data is a list
    if isinstance(json_data, list):
        # Convert list to dict with 'items' key
        json_data = {'items': json_data}
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    
    # Set sheet name based on document type if provided
    if doc_type_code:
        ws.title = f"{doc_type_code[:24]}"  # Excel sheet names are limited to 31 chars
    else:
        ws.title = "Data"
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    row_num = 1
    
    # Process metadata (non-array fields) first
    metadata = {}
    tables = {}
    
    for key, value in json_data.items():
        if isinstance(value, list) and all(isinstance(x, dict) for x in value):
            # This is a table - save for later
            tables[key] = value
        elif isinstance(value, dict):
            # Flatten nested dictionaries
            flattened = flatten_dict(value)
            for sub_key, sub_value in flattened.items():
                full_key = f"{key}_{sub_key}" if sub_key else key
                metadata[full_key] = sub_value
        else:
            # Simple key-value
            metadata[key] = value
    
    # Write metadata section if we have any
    if metadata:
        # Write section header
        ws.cell(row=row_num, column=1, value="Document Information")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        row_num += 1
        
        # Write headers
        ws.cell(row=row_num, column=1, value="Field").font = header_font
        ws.cell(row=row_num, column=2, value="Value").font = header_font
        ws.cell(row=row_num, column=1).fill = header_fill
        ws.cell(row=row_num, column=2).fill = header_fill
        row_num += 1
        
        # Write metadata
        for key, value in metadata.items():
            display_key = format_key_for_display(key)
            ws.cell(row=row_num, column=1, value=display_key)
            ws.cell(row=row_num, column=2, value=str(value))
            row_num += 1
            
        # Add a blank row
        row_num += 1
    
    # Process each table
    for table_name, table_data in tables.items():
        if not table_data:
            continue
            
        # Write table header
        display_table_name = format_key_for_display(table_name)
        ws.cell(row=row_num, column=1, value=display_table_name)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        row_num += 1
        
        # Get all unique keys from all objects in the array
        all_keys = set()
        for item in table_data:
            # For each object, get flattened keys
            flattened_item = flatten_dict(item)
            all_keys.update(flattened_item.keys())
            
        # Sort keys for consistent order
        sorted_keys = sorted(all_keys)
        
        # Write table headers
        col_num = 1
        for key in sorted_keys:
            display_key = format_key_for_display(key)
            cell = ws.cell(row=row_num, column=col_num, value=display_key)
            cell.font = header_font
            cell.fill = header_fill
            col_num += 1
        row_num += 1
        
        # Write table rows
        for item in table_data:
            flattened_item = flatten_dict(item)
            col_num = 1
            for key in sorted_keys:
                value = flattened_item.get(key, '')
                ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
                col_num += 1
            row_num += 1
            
        # Add a blank row after the table
        row_num += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width at 50
    
    # Save the workbook
    wb.save(output_path)
    return output_path 