from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import os
from werkzeug.utils import secure_filename
import json
from datetime import datetime
import sys
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import functions from main.py
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from main import extract_text_from_image, configure_enhanced_prompt, get_response_schema

app = Flask(__name__)
CORS(app)

# Update these lines near the top of app.py
UPLOAD_FOLDER = "uploads"
JSON_OUTPUT_FOLDER = "output/json"
EXCEL_OUTPUT_FOLDER = "output/excel"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB max upload

# Create all necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(JSON_OUTPUT_FOLDER, exist_ok=True)
os.makedirs(EXCEL_OUTPUT_FOLDER, exist_ok=True)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process_image():
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files["file"]
    invoice_type = request.form.get("invoice_type", "printing")

    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        upload_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(upload_path)

        try:
            # Get API key from environment
            API_KEY = "AIzaSyDnUHmWDSsrh3X6wImAH4UOgRV1kLUA41E"

            # Process image using existing functions
            enhanced_prompt = configure_enhanced_prompt(invoice_type)
            response_schema = get_response_schema(invoice_type)
            extracted_text = extract_text_from_image(
                upload_path, enhanced_prompt, response_schema, API_KEY
            )

            # Save results
            output_filename = f"{invoice_type}_{timestamp}.json"
            output_path = os.path.join(JSON_OUTPUT_FOLDER, output_filename)

            try:
                json_data = json.loads(extracted_text)
                with open(output_path, "w", encoding="utf-8") as json_file:
                    json.dump(json_data, json_file, indent=2, ensure_ascii=False)
            except json.JSONDecodeError:
                with open(output_path, "w", encoding="utf-8") as json_file:
                    json.dump(
                        {"raw_text": extracted_text},
                        json_file,
                        indent=2,
                        ensure_ascii=False,
                    )

            return jsonify(
                {
                    "success": True,
                    "filename": output_filename,
                    "results": (
                        json.loads(extracted_text)
                        if isinstance(extracted_text, str)
                        else extracted_text
                    ),
                }
            )

        except Exception as e:
            return jsonify({"error": str(e)}), 500

    return jsonify({"error": "Invalid file type"}), 400


@app.route("/download/<filename>")
def download_file(filename):
    return send_file(os.path.join(JSON_OUTPUT_FOLDER, filename), as_attachment=True)


@app.route("/download-excel/<filename>")
def download_excel(filename):
    json_path = os.path.join(JSON_OUTPUT_FOLDER, filename)
    try:
        # Read the JSON file
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Extract invoice type from filename
        # The filename format is {invoice_type}_{timestamp}.json
        invoice_type = filename.split("_")[0]

        if invoice_type == "shop":
            # Define headers
            headers = [
                "No",
                "Item Name",
                "Quantity",
                "Unit Price",
                "Amount",
                "Shop Address",
                "Shop Code",
                "Shop Telephone",
                "Shop Name",
                "Issue DateTime",
                "Pos Terminal",
                "Issue Number",
                "Brand",
                "Payment",
                "Remark",
                "Subtotal",
                "Total Amount",
                "Discount",
            ]

            # Prepare data rows (without headers)
            data_rows = []

            for index, row in enumerate(data):
                issue_info = row["issue_info"]
                shop_address = issue_info.get("shop_address", "N/A")
                shop_code = issue_info.get("shop_code", "N/A")
                shop_telephone = issue_info.get("shop_telephone", "N/A")
                shop_name = issue_info.get("brand", "N/A")
                issue_datetime = issue_info.get("issue_datetime", "N/A")
                pos_terminal = issue_info.get("pos_terminal", "N/A")
                issue_number = issue_info.get("issue_number", "N/A")
                brand = issue_info.get("brand", "N/A")
                payment = row.get("payment", "N/A")
                remark = row.get("remark", "N/A")
                subtotal = row.get("subtotal", "N/A")
                total_amount = row.get("total_amount", "N/A")
                discount = row.get("discount", "N/A")
                details = row.get("details", "N/A")
                for detail in details:
                    item_name = detail.get("item_name", "N/A")
                    quantity = detail.get("quantity", "N/A")
                    unit_price = detail.get("unit_price", "N/A")
                    amount = detail.get("amount", "N/A")
                    data_row = [
                        index + 1,
                        item_name,
                        quantity,
                        unit_price,
                        amount,
                        shop_address,
                        shop_code,
                        shop_telephone,
                        shop_name,
                        issue_datetime,
                        pos_terminal,
                        issue_number,
                        brand,
                        payment,
                        remark,
                        subtotal,
                        total_amount,
                        discount,
                    ]
                    data_rows.append(data_row)

        elif invoice_type == "printing":
            # Define headers for printing invoice
            headers = [
                "No",
                "Date",
                "Reference",
                "Description",
                "Quantity",
                "Amount",
                "Issuer Company",
                "Issuer Address",
                "Issuer Phone",
                "Receiver Company",
                "Receiver ID",
                "Receiver Address",
                "Receiver Contact Person",
                "Receiver Fax",
                "Recipient",
                "Recipient Contact",
                "Service Period",
                "Statement Date",
                "Currency",
                "Total",
            ]

            # Prepare data rows (without headers)
            data_rows = []

            for index, row in enumerate(data):
                issuer_info = row.get("issuer_info", {})
                receiver_info = row.get("receiver_info", {})

                # Extract issuer information
                issuer_company = issuer_info.get("company_name", {}).get("text", "N/A")
                issuer_address = issuer_info.get("company_address", {}).get(
                    "text", "N/A"
                )
                issuer_phone = issuer_info.get("contact_phone", {}).get("text", "N/A")

                # Extract receiver information
                receiver_company = receiver_info.get("company_name", {}).get(
                    "text", "N/A"
                )
                receiver_id = receiver_info.get("company_id", {}).get("text", "N/A")
                receiver_address = receiver_info.get("company_address", {}).get(
                    "text", "N/A"
                )
                receiver_contact = receiver_info.get("contact_person", {}).get(
                    "text", "N/A"
                )
                receiver_fax = receiver_info.get("fax", {}).get("text", "N/A")
                recipient = receiver_info.get("recipient", {}).get("text", "N/A")
                recipient_contact = receiver_info.get("recipient_contact", {}).get(
                    "text", "N/A"
                )

                # Extract invoice metadata
                service_period = row.get("service_period", {}).get("text", "N/A")
                # invoice_number = row.get('invoice_number', {}).get('text', 'N/A')
                statement_date = row.get("statement_date", {}).get("text", "N/A")
                currency = row.get("currency", {}).get("text", "N/A")
                # subtotal = row.get('subtotal', {}).get('text', 'N/A')
                total = row.get("total_amount", {}).get("text", "N/A")
                # terms = row.get('terms', {}).get('text', 'N/A')

                # Process each detail row
                details = row.get("details", [])
                for detail in details:
                    date = detail.get("date", {}).get("text", "N/A")
                    ref = detail.get("ref", {}).get("text", "N/A")
                    desc = detail.get("desc", {}).get("text", "N/A")
                    quantity = detail.get("quantity", {}).get("text", "N/A")
                    amount = detail.get("amt", {}).get("text", "N/A")

                    data_row = [
                        index + 1,
                        date,
                        ref,
                        desc,
                        quantity,
                        amount,
                        issuer_company,
                        issuer_address,
                        issuer_phone,
                        receiver_company,
                        receiver_id,
                        receiver_address,
                        receiver_contact,
                        receiver_fax,
                        recipient,
                        recipient_contact,
                        service_period,
                        statement_date,
                        currency,
                        total,
                    ]
                    data_rows.append(data_row)
        else:
            return jsonify({"error": f"Unknown invoice type: {invoice_type}"}), 400

        # Create DataFrame with proper headers
        df = pd.DataFrame(data_rows, columns=headers)

        # Create Excel file
        excel_filename = filename.replace(".json", ".xlsx")
        excel_path = os.path.join(EXCEL_OUTPUT_FOLDER, excel_filename)

        # Export with formatting
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Invoice Details", index=False)

            # Format the Excel file
            workbook = writer.book
            worksheet = writer.sheets["Invoice Details"]

            # Format headers
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center")

            # Add headers styling
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = header_alignment

            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width

        return send_file(excel_path, as_attachment=True, download_name=excel_filename)
    except Exception as e:
        import traceback

        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, ssl_context="adhoc", debug=True)
