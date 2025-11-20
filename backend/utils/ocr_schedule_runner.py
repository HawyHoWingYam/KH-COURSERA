"""Background OCR schedule runner.

This module implements the recurring logic for:
- Selecting due OCR schedules
- Ensuring monthly OneDrive folder / Excel structure
- Discovering new material files
- Running OCR and appending results to Excel
- Moving processed files to history and tracking status in DB
"""

import json
import logging
import os
import tempfile
from datetime import datetime, time, timedelta
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session
from sqlalchemy import and_

from db.database import SessionLocal
from db.models import (
    OcrSchedule,
    OcrScheduledFile,
    ScheduleMode,
    ScheduledFileStatus,
)
from main import extract_text_from_pdf
from utils.order_processor import escape_excel_formulas
from utils.onedrive_client import (
    build_client_from_env,
    normalise_onedrive_path,
    join_onedrive_path,
)

logger = logging.getLogger(__name__)


def _utcnow() -> datetime:
    return datetime.utcnow()


def _parse_time_hhmm(value: Optional[str]) -> Optional[time]:
    """Parse HH:MM string to time, returning None on invalid input."""
    if not value:
        return None
    try:
        parts = value.split(":")
        if len(parts) != 2:
            return None
        hour = int(parts[0])
        minute = int(parts[1])
        return time(hour=hour, minute=minute)
    except Exception:
        return None


def _is_within_window(schedule: OcrSchedule, now: datetime) -> bool:
    """Check whether 'now' is inside the schedule's allowed window."""
    if schedule.schedule_mode != ScheduleMode.WINDOWED_INTERVAL:
        return True

    # Weekday filtering
    if schedule.allowed_weekdays:
        try:
            allowed = {
                int(token)
                for token in schedule.allowed_weekdays.split(",")
                if token.strip()
            }
        except Exception:
            allowed = set()
        if allowed and now.weekday() not in allowed:
            return False

    # Time-of-day window (treated in server local time for simplicity)
    start_t = _parse_time_hhmm(schedule.window_start_time)
    end_t = _parse_time_hhmm(schedule.window_end_time)

    if start_t is None and end_t is None:
        return True

    local_t = now.time()

    if start_t and end_t:
        # Simple inclusive start, exclusive end window
        if start_t <= end_t:
            return start_t <= local_t < end_t
        # Overnight window (e.g. 22:00–06:00)
        return local_t >= start_t or local_t < end_t
    if start_t:
        return local_t >= start_t
    if end_t:
        return local_t < end_t

    return True


def _get_current_month_str(now: datetime) -> str:
    return now.strftime("%Y%m")


def _ensure_month_structure(
    onedrive_client,
    schedule: OcrSchedule,
    month_str: str,
) -> Tuple[Any, Any, Any, str]:
    """Ensure Material/YYYYMM, History/YYYYMM and Output/YYYYMM.xlsx exist."""
    # Material root
    material_root = normalise_onedrive_path(schedule.material_root_path or "")
    history_root = normalise_onedrive_path(schedule.history_root_path or "")
    output_root = normalise_onedrive_path(schedule.output_root_path or "")

    if not material_root or not history_root or not output_root:
        raise RuntimeError(
            f"OCR schedule {schedule.schedule_id} has incomplete OneDrive roots"
        )

    material_parent = onedrive_client.get_folder(material_root)
    if not material_parent:
        raise RuntimeError(
            f"Material root folder not found for schedule {schedule.schedule_id}: {material_root}"
        )

    history_parent = onedrive_client.get_folder(history_root)
    if not history_parent:
        raise RuntimeError(
            f"History root folder not found for schedule {schedule.schedule_id}: {history_root}"
        )

    output_parent = onedrive_client.get_folder(output_root)
    if not output_parent:
        raise RuntimeError(
            f"Output root folder not found for schedule {schedule.schedule_id}: {output_root}"
        )

    material_month = onedrive_client.get_or_create_folder(material_parent, month_str)
    if not material_month:
        raise RuntimeError(
            f"Failed to get/create material month folder {month_str} under {material_root}"
        )

    failed_name = schedule.failed_subfolder_name or "_Failed"
    failed_folder = onedrive_client.get_or_create_folder(material_month, failed_name)
    if not failed_folder:
        raise RuntimeError(
            f"Failed to get/create failed folder '{failed_name}' under material month {month_str}"
        )

    history_month = onedrive_client.get_or_create_folder(history_parent, month_str)
    if not history_month:
        raise RuntimeError(
            f"Failed to get/create history month folder {month_str} under {history_root}"
        )

    # Excel file path is just a canonical OneDrive path; the actual upload
    # is handled lazily when we first need to write.
    output_excel_path = join_onedrive_path(output_root, f"{month_str}.xlsx")

    return material_month, failed_folder, history_month, output_excel_path


def _discover_candidates(
    db: Session,
    schedule: OcrSchedule,
    month_str: str,
    onedrive_client,
    material_month_folder,
    max_files: int,
) -> List[Tuple[Any, OcrScheduledFile]]:
    """Discover new or retryable files in the monthly material folder."""
    candidates: List[Tuple[Any, OcrScheduledFile]] = []

    # For now, reuse list_all_pdfs and then filter on name.
    # Month filtering is implicit via the folder name (YYYYMM).
    from utils.onedrive_client import O365File  # type: ignore

    try:
        pdf_items: List[O365File] = onedrive_client.list_all_pdfs(
            material_month_folder,
            created_month_filter=None,
        )
    except Exception as exc:
        logger.error(
            "❌ Failed to list PDFs in material folder for schedule %s: %s",
            schedule.schedule_id,
            exc,
        )
        return []

    for item in pdf_items:
        name = getattr(item, "name", None) or ""
        if not name or name.startswith("~$"):
            # Skip temporary lock files
            continue

        onedrive_path = join_onedrive_path(
            schedule.material_root_path or "",
            month_str,
            name,
        )

        row: Optional[OcrScheduledFile] = (
            db.query(OcrScheduledFile)
            .filter(
                and_(
                    OcrScheduledFile.schedule_id == schedule.schedule_id,
                    OcrScheduledFile.onedrive_path == onedrive_path,
                )
            )
            .one_or_none()
        )

        if row:
            if row.status in (
                ScheduledFileStatus.COMPLETED,
                ScheduledFileStatus.PROCESSING,
            ):
                continue
            # Allow retries for WAITING_FOR_EXCEL or ERROR
        else:
            row = OcrScheduledFile(
                schedule_id=schedule.schedule_id,
                month_str=month_str,
                onedrive_path=onedrive_path,
                filename=name,
                status=ScheduledFileStatus.PENDING,
                attempt_count=0,
            )
            db.add(row)
            db.flush()

        candidates.append((item, row))

        if max_files and len(candidates) >= max_files:
            break

    db.commit()
    return candidates


def _append_rows_to_excel_on_onedrive(
    onedrive_client,
    excel_path: str,
    ocr_json: Dict[str, Any],
    max_retries: int = 3,
) -> int:
    """Append flattened OCR JSON rows to Excel file on OneDrive, handling lock/retry.

    Uses the same deep-flattening strategy as json_to_csv/deep_flatten_json_universal so
    that structure is driven by the document schema and Gemini output, not by a
    per-schedule Excel configuration.

    Returns the 1-based Excel row index of the first appended row, or -1 on failure.
    """
    import io

    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover - environment guard
        logger.error("❌ openpyxl is required for Excel append but not installed: %s", exc)
        return -1

    # Reuse the same deep flattening logic used for consolidated CSV/Excel
    from utils.excel_converter import deep_flatten_json_universal

    flattened_rows: List[Dict[str, Any]] = deep_flatten_json_universal(ocr_json)
    if not flattened_rows:
        logger.warning("⚠️ No data found in OCR JSON to append to Excel")
        return -1

    # Determine header set for new data
    new_headers: List[str] = []
    seen = set()
    for row in flattened_rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                new_headers.append(str(key))

    last_exc: Optional[Exception] = None
    for attempt in range(max_retries):
        try:
            # Download existing workbook if present; otherwise create new one.
            content = onedrive_client.download_file_content(excel_path)
            if content:
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                # Existing header from first row
                existing_headers = [
                    cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                header_cols = [str(h) for h in existing_headers]
                # NOTE: For now we do NOT auto-extend headers if new keys appear later.
                # This keeps the sheet schema stable once created.
                start_row_index = ws.max_row + 1
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                # First-time header: union of keys across flattened rows
                header_cols = new_headers
                ws.append(header_cols)
                start_row_index = 2

            # Append one Excel row per flattened record
            for record in flattened_rows:
                row_values: List[Any] = []
                for col_name in header_cols:
                    raw_value = record.get(col_name)
                    safe_value = escape_excel_formulas(raw_value)
                    row_values.append(safe_value)
                ws.append(row_values)

            # Save workbook back to bytes
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            data = bio.read()

            # Upload back to same path
            normalized_path = normalise_onedrive_path(excel_path)
            file_item = onedrive_client.drive.get_item_by_path(normalized_path)
            if file_item:
                # Replace existing content
                file_item.upload(data)
            else:
                # Create a new file under the parent folder
                parent_path = "/".join(normalized_path.split("/")[:-1])
                parent_folder = onedrive_client.get_folder(parent_path)
                if not parent_folder:
                    logger.error(
                        "❌ Failed to resolve parent folder %s when uploading Excel",
                        parent_path,
                    )
                    return -1
                parent_folder.upload_file(
                    data=data,
                    name=normalized_path.split("/")[-1],
                )

            logger.info("✅ Appended row to Excel at %s (row %s)", excel_path, start_row_index)
            return start_row_index

        except Exception as exc:
            last_exc = exc
            logger.warning(
                "⚠️ Excel append attempt %s/%s failed for %s: %s",
                attempt + 1,
                max_retries,
                excel_path,
                exc,
            )
            if attempt < max_retries - 1:
                # Patient retry strategy: 5s, then 10s, etc.
                delay = 5 * (attempt + 1)
                logger.info("⏳ Waiting %s seconds before retrying Excel append", delay)
                from time import sleep

                sleep(delay)

    if last_exc:
        logger.error("❌ All Excel append attempts failed for %s: %s", excel_path, last_exc)
    return -1


def _process_single_file(
    db: Session,
    schedule: OcrSchedule,
    onedrive_client,
    file_item,
    row: OcrScheduledFile,
    month_str: str,
    history_month_folder,
    failed_folder,
    output_excel_path: str,
) -> None:
    """Run full pipeline for a single file."""
    now = _utcnow()

    # Lock row
    row = (
        db.query(OcrScheduledFile)
        .filter(OcrScheduledFile.id == row.id)
        .with_for_update()
        .one()
    )
    if row.status in (
        ScheduledFileStatus.PROCESSING,
        ScheduledFileStatus.COMPLETED,
    ):
        return

    row.status = ScheduledFileStatus.PROCESSING
    row.attempt_count = (row.attempt_count or 0) + 1
    row.error_message = None
    db.commit()

    # Download file to temp dir
    filename = row.filename or getattr(file_item, "name", "file.pdf")
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            ok = onedrive_client.download_file(file_item, tmpdir)
            if not ok:
                raise RuntimeError("Download from OneDrive failed")

            local_path = os.path.join(tmpdir, filename)

            # Raw OCR -> JSON
            ocr_result = extract_text_from_pdf(local_path)
    except Exception as exc:
        logger.error(
            "❌ OCR failed for schedule %s file %s: %s",
            schedule.schedule_id,
            filename,
            exc,
        )
        # Move to failed folder
        try:
            onedrive_client.move_file(file_item, failed_folder)
        except Exception as move_exc:
            logger.error(
                "❌ Additionally failed to move file %s to failed folder: %s",
                filename,
                move_exc,
            )
        row.status = ScheduledFileStatus.ERROR
        row.error_message = f"OCR failed: {exc}"
        db.commit()
        return

    # Persist OCR JSON if desired (optional; for now just store inline as JSON string)
    try:
        row.ocr_json_path = json.dumps(ocr_result, ensure_ascii=False)
    except Exception:
        # Best-effort; keep going even if we cannot serialize
        row.ocr_json_path = None

    # Excel append (with lock-aware retries)
    excel_row_index = _append_rows_to_excel_on_onedrive(
        onedrive_client,
        output_excel_path,
        ocr_result or {},
    )
    if excel_row_index <= 0:
        row.status = ScheduledFileStatus.WAITING_FOR_EXCEL
        row.error_message = (
            row.error_message or "Excel locked or write failed; will retry later"
        )
        db.commit()
        return

    # Move file to history/YYYYMM
    from O365.drive import File as O365File, Folder as O365Folder  # type: ignore

    try:
        existing_names = set()
        for item in history_month_folder.get_items():
            if getattr(item, "is_file", False):
                existing_names.add(getattr(item, "name", ""))

        target_name = filename
        if target_name in existing_names:
            # Add timestamp suffix to avoid collision
            stem, dot, suffix = target_name.partition(".")
            ts = now.strftime("%Y%m%d_%H%M%S")
            target_name = f"{stem}_{ts}.{suffix}" if dot else f"{stem}_{ts}"

        onedrive_client.move_file(file_item, history_month_folder, new_name=target_name)
    except Exception as exc:
        logger.error(
            "❌ Failed to move file %s to history for schedule %s: %s",
            filename,
            schedule.schedule_id,
            exc,
        )
        # We still consider OCR+Excel successful; leave file in place for manual intervention.

    # Finalize
    row.status = ScheduledFileStatus.COMPLETED
    row.output_excel_path = output_excel_path
    row.excel_row_index = excel_row_index
    db.commit()


def process_schedule(schedule_id: int) -> None:
    """Process a single OCR schedule if its OneDrive roots and config are valid."""
    db: Session = SessionLocal()
    try:
        schedule = db.query(OcrSchedule).filter(
            OcrSchedule.schedule_id == schedule_id
        ).one_or_none()
        if not schedule:
            logger.warning("⚠️ OCR schedule %s not found", schedule_id)
            return
        if not schedule.enabled:
            logger.info("ℹ️ OCR schedule %s is disabled; skipping", schedule_id)
            return

        now = _utcnow()
        month_str = _get_current_month_str(now)

        onedrive_client = build_client_from_env()
        if not onedrive_client:
            logger.error("❌ Cannot build OneDrive client from env; aborting schedule %s", schedule_id)
            return
        if not onedrive_client.connect():
            logger.error("❌ Failed to connect to OneDrive for schedule %s", schedule_id)
            return

        material_month, failed_folder, history_month, output_excel_path = _ensure_month_structure(
            onedrive_client,
            schedule,
            month_str,
        )

        max_files = getattr(schedule, "max_files_per_cycle", None)
        try:
            max_files_int = int(max_files) if max_files is not None else 10
        except Exception:
            max_files_int = 10

        candidates = _discover_candidates(
            db=db,
            schedule=schedule,
            month_str=month_str,
            onedrive_client=onedrive_client,
            material_month_folder=material_month,
            max_files=max_files_int,
        )

        if not candidates:
            logger.info("ℹ️ No OCR candidates found for schedule %s", schedule_id)
            return

        logger.info(
            "🔄 Processing %s OCR files for schedule %s (month %s)",
            len(candidates),
            schedule_id,
            month_str,
        )

        for file_item, row in candidates:
            _process_single_file(
                db=db,
                schedule=schedule,
                onedrive_client=onedrive_client,
                file_item=file_item,
                row=row,
                month_str=month_str,
                history_month_folder=history_month,
                failed_folder=failed_folder,
                output_excel_path=output_excel_path,
            )
    finally:
        try:
            db.close()
        except Exception:
            pass


def run_ocr_schedules() -> None:
    """Entry point for APScheduler.

    This function selects all due schedules and processes each in turn.
    """
    db: Session = SessionLocal()
    try:
        now = _utcnow()

        due_schedules: List[OcrSchedule] = (
            db.query(OcrSchedule)
            .filter(
                and_(
                    OcrSchedule.enabled.is_(True),
                    OcrSchedule.next_run_at <= now,
                )
            )
            .all()
        )

        if not due_schedules:
            return

        logger.info("🔄 Found %s due OCR schedules", len(due_schedules))

        for schedule in due_schedules:
            if schedule.start_at and now < schedule.start_at:
                continue
            if not _is_within_window(schedule, now):
                logger.info(
                    "⏸️ Schedule %s is outside configured window; skipping this cycle",
                    schedule.schedule_id,
                )
                # Push next_run_at forward to avoid hammering
                schedule.next_run_at = now + timedelta(
                    seconds=schedule.interval_seconds or 300
                )
                db.commit()
                continue

            # Advance next_run_at before processing to avoid re-selection
            schedule.last_run_at = now
            schedule.next_run_at = now + timedelta(
                seconds=schedule.interval_seconds or 300
            )
            db.commit()

            try:
                process_schedule(schedule.schedule_id)
            except Exception as exc:
                logger.error(
                    "❌ Error while processing OCR schedule %s: %s",
                    schedule.schedule_id,
                    exc,
                )
    finally:
        try:
            db.close()
        except Exception:
            pass
